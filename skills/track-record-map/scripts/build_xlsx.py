import os
WORKDIR = os.environ.get('TRM_WORKDIR', os.getcwd())

# -*- coding: utf-8 -*-
import json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE = WORKDIR
MERGED_XLSX_SRC = os.path.join(BASE, 'merged_for_xlsx.json')
CACHE_PATH = os.path.join(BASE, 'geocode_cache.json')
OUT_XLSX = os.environ.get('TRM_XLSX_OUT', os.path.join(WORKDIR, 'track-record-export.xlsx'))

with open(MERGED_XLSX_SRC, encoding='utf-8') as f:
    records = json.load(f)

with open(CACHE_PATH, encoding='utf-8') as f:
    cache = json.load(f)

def cache_key(addr, city):
    addr_no_unit = re.sub(r',?\s*#\S+$', '', addr).strip()
    return f'{addr_no_unit}|{city}'

geocoded_ok = 0
geocoded_fail = 0
for r in records:
    k = cache_key(r['address'], r['city'])
    v = cache.get(k)
    if v:
        r['lat'], r['lon'] = v[0], v[1]
        geocoded_ok += 1
    else:
        r['lat'], r['lon'] = None, None
        geocoded_fail += 1

def parse_date(d):
    m, day, y = d.split('/')
    return (int(y), int(m), int(day))

records.sort(key=lambda r: parse_date(r['date']))

wb = Workbook()
ws = wb.active
ws.title = os.environ.get('TRM_SHEET_TITLE', 'Sold Track Record')[:31]

headers = ['MLS #', 'Address', 'City', 'Sale Price', 'Close Date', 'Property Type',
           'Beds', 'Baths (Full|Half)', 'SqFt', 'Lot Size', 'Matched DRE / Side(s)',
           'Latitude', 'Longitude']

header_fill = PatternFill(start_color='1C2B45', end_color='1C2B45', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for i, r in enumerate(records, start=2):
    m, day, y = r['date'].split('/')
    close_date_str = f'{y}-{m}-{day}'
    ws.cell(row=i, column=1, value=r['mls'])
    ws.cell(row=i, column=2, value=r['address'])
    ws.cell(row=i, column=3, value=r['city'])
    price_cell = ws.cell(row=i, column=4, value=r['price'])
    price_cell.number_format = '$#,##0'
    date_cell = ws.cell(row=i, column=5, value=close_date_str)
    ws.cell(row=i, column=6, value=r['cls'])
    ws.cell(row=i, column=7, value=r['beds'] if r['beds'] else None)
    ws.cell(row=i, column=8, value=r['baths'] if r['baths'] else None)
    ws.cell(row=i, column=9, value=r['sqft'].replace(',', '') if r['sqft'] else None)
    ws.cell(row=i, column=10, value=r['lot_size'] if r['lot_size'] else None)
    ws.cell(row=i, column=11, value='; '.join(r['sources']))
    ws.cell(row=i, column=12, value=r['lat'])
    ws.cell(row=i, column=13, value=r['lon'])

widths = [13, 30, 16, 13, 12, 14, 7, 15, 8, 16, 46, 11, 11]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(records)+1}'

os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
wb.save(OUT_XLSX)

print(f'Wrote {len(records)} rows to {OUT_XLSX}')
print(f'Geocoded (lat/lon present): {geocoded_ok}')
print(f'Missing lat/lon: {geocoded_fail}')
